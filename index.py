import openpyxl as xl
wb = xl.load_workbook("transactions.xlsx")
sheet = wb['Sheet1']
# cell = sheet['a1']
# cell = sheet.cell(1, 1)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_Value = cell.value * 0.9
    corrected_Value_cell = sheet.cell(row, 4)
    corrected_Value_cell.value = corrected_Value
    print(sheet.cell(row, 4).value)

wb.save("transactions2.xlsx")
